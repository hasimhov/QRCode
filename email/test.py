# import base64 
import os
# import requests
import smtplib
# import json
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart	
import getpass
import openpyxl
# import lxml.html
# import openpyxl
# import codecs
import imghdr
from email.utils import make_msgid
import csv
from email.message import EmailMessage


# def encode(image):
# 	image_64_encode = base64.encodestring(image)
# 	return image_64_encode

if __name__ == "__main__":
	s = smtplib.SMTP()
	# s.connect('smtp.sendgrid.net', '587')
	s.connect('smtp.gmail.com')
	s.ehlo()
	s.starttls()
	s.ehlo()
	# sent=[]
	# excel = openpyxl.load_workbook(filename = "Sent.xlsx")
	# for sheet in excel.get_sheet_names():
	# 	sheetdata = excel[sheet]
	# 	i=1
	# 	data=sheetdata.cell(row=i,column=1).value
	# 	print(data)
	# 	while data not in [None,''] and ('@smail.iitm.ac.in' in data):
	# 		sent.append(data)
	# 		i=i+1
	# 		data=sheetdata.cell(row=i,column=1).value
	# 	break

	# print(sent)
	# Password = 'Mahesh!1'
	Password = getpass.getpass()
	s.login("hsehamtk2299@gmail.com",Password)
	# imageopen = open('saarang.jpg','rb')
	# image_read = imageopen.read()
	# encodedimage1 = encode(image_read)

	# imageopen = open('Choreo_Night.jpeg','rb')
	# image_read = imageopen.read()
	# encodedimage2 = encode(image_read)

	# imageopen = open('EDM_Night.jpg','rb')
	# image_read = imageopen.read()
	# encodedimage3 = encode(image_read)

	# imageopen = open('Popular_Night.jpg','rb')
	# image_read = imageopen.read()
	# encodedimage4 = encode(image_read)


	# rock_show_file = open('list_rock.csv','rt')
	# edm_night_file = open('list_edm.csv','rt')
	# choreo_night_file = open('list_choreo.csv','rt')
	# popular_night_file = open('list_popular.csv','rt')


	# rock_show = csv.reader(rock_show_file)
	# edm_night = csv.reader(edm_night_file)
	# choreo_night = csv.reader(choreo_night_file)
	# popular_night = csv.reader(popular_night_file)

	# rock_show_list = []
	# edm_night_list = []
	# choreo_night_list = []
	# popular_night_list = []

	# for row in rock_show:
	# 	if row[2] == 'Student Id':
	# 		continue
	# 	rock_show_list.append(row[2].upper())
	
	# for row in edm_night:
	# 	if row[2] == 'Student Id':
	# 		continue
	# 	edm_night_list.append(row[2].upper())

	# for row in choreo_night:
	# 	if row[2] == 'Student Id':
	# 		continue
	# 	choreo_night_list.append(row[2].upper())

	# for row in popular_night:
	# 	if row[2] == 'Student Id':
	# 		continue
	# 	popular_night_list.append(row[2].upper())

	# # encodedimage5 = encode(image_read)
	# i = 1
	# # imageopen.close()
	# for filename in os.listdir('Attachments/'):
	# # for filename in ['CS16B045.png']:
	# 	rollno,ext = os.path.splitext(filename)

	msg=EmailMessage()
	asparagus_cid = make_msgid()
	asparagus_cid1 = make_msgid()
	# asparagus_cid2 = make_msgid()
	# asparagus_cid3 = make_msgid()
	# asparagus_cid4 = make_msgid()
	# QRCode = make_msgid()
	msg.set_content("")
	msg.add_alternative("""\
	<html>
	  <body>
	  	<div style="text-align:center;">
	  		<img src="cid:{asparagus_cid}" style="width:300px;" \>
	  	</div>
	    <p>Greetings from Saarang 2018. Your QR code has been sent to your smail inbox. 
	    The mail was sent from sec_lit@smail.iitm.ac.in with the subject:<strong> <Your roll no.> - Your QR code for Saarang 2018 Proshows.</strong></p>
		<p><strong>Mess card not applicable.</strong></p><br>
		<p>Search for the subject mentioned above in the inbox or if you still have trouble finding it, check your promotions/spam folder. 
		Also, the mail has been sent to you between <srong>12:00 AM - 3:00 AM on 10th January 2018</strong></p>
		<p>The instructions for entering into OAT via the QR code are as follows:</p>
	  	<div style="text-align:center;">
	    	<img src="cid:{asparagus_cid1}" style="width:500px;" \>
	    </div>
	    <p>You can present the mail/downloaded attachment/printout of the above mentioned received mail attachment for entrance.</p>
	    <p>You have to use the <strong>same QR code</strong> for all the shows that you have bought tickets for.</p><br>
		<p>Please note that the entry is valid <string>only on your institute ID card (Mess card not applicable).</strong>
		In case you have lost your Institute ID card or are unable to present it, collect your ticket from <strong>Faculty Association Office, 
		MSB</strong> from <strong>11:00 AM to 1:00 PM</strong> on the day of the show.
		<strong>Photo ID other than mess card</strong> is required to collect your ticket.</p><br>
		<p>In case you have not receive the mail or there is someone else’s roll number mentioned in the attachment or the mail, please 
		contact us as soon as possible. Also, read the instructions given in that mail carefully.</p><br>
		<p>For any queries related to this please contact sales@saarang.org</p>		
	  </body>
	</html>
	""".format(asparagus_cid=asparagus_cid[1:-1],asparagus_cid1=asparagus_cid1[1:-1]), subtype='html')
	# note that we needed to peel the <> off the msgid for use in the html.

	

	# Now add the related image to the html part.
	with open("images/image2.jpg",'rb') as img:
	    msg.get_payload()[1].add_related(img.read(), 'image', 'png',cid=asparagus_cid)

	with open("images/image1.jpg",'rb') as img:
	    msg.get_payload()[1].add_related(img.read(), 'image', 'png',cid=asparagus_cid1)

		# if rollno in choreo_night_list:
		# 	with open("Choreo_Night.jpeg", 'rb') as img:
		# 	    msg.get_payload()[1].add_related(img.read(), 'image', 'jpeg',cid=asparagus_cid1)

		# if rollno in edm_night_list:
		# 	with open("EDM_Night.jpg", 'rb') as img:
		# 	    msg.get_payload()[1].add_related(img.read(), 'image', 'jpg',cid=asparagus_cid2)

		# if rollno in rock_show_list:
		# 	with open("Rock_Show.jpeg", 'rb') as img:
		# 	    msg.get_payload()[1].add_related(img.read(), 'image', 'jpeg',cid=asparagus_cid3)

		# if rollno in popular_night_list:
		# 	with open("Popular_Night.jpg", 'rb') as img:
		# 	    msg.get_payload()[1].add_related(img.read(), 'image', 'jpg',cid=asparagus_cid4)

		# image_read = open('Attachments/'+filename,'rb').read()
		# msg.add_attachment(image_read,maintype='image',subtype=imghdr.what(None,image_read),filename='SaarangQR.jpg')
	msg['From'] = 'hsehamtk2299@gmail.com'
	msg['Subject'] = 'QR code for Proshow tickets bought on iKollege'
	msg['To'] = 'divyanshup@saarang.org'
		# msg['From']=str(rollno).lower()+'@saarang.org'
		# msg['To']=['divyanshup@saarang.org','rohit@saarang.org','hsehamtk2299@gmail.com','rajat_kumar@saarang.org','navaneeth@saarang.org','anikesh@saarang.org']
		# msg['To'] = 'ch15b061@smail.iitm.ac.in'
		# msg['To'] = str(rollno).lower()+'@smail.iitm.ac.in'
		# print(msg['To'],end='\t')
		# print(i)
		# i+=1
	s.sendmail(msg['From'],msg['To'],msg.as_string())
		# os.remove('Attachments/'+str(rollno).upper()+'.jpg')

	# print(sent)

	# for filename in os.listdir('QRCodes/Boys1/'):
	# 	msg = MIMEMultipart()
	# 	msg['From'] =  'do-not-reply@saarang.org'
	# 	body, ext = os.path.splitext(filename)
	# 	emailId=str.lower(str(body))+"@smail.iitm.ac.in"
	# 	imageopen = open(os.path.join('QRCodes/Boys1/',filename), 'rb')
	# 	image_read = imageopen.read()
	# 	# encodedimage = encode(image_read)
	# 	# print(str(encodedimage))
	# 	# print(emailId)
	# 	# print(type(msg['To']))
	# 	# msg['To'] = emailId
	# 	msg['To'] = emailId
	# 	print(msg['To'])
	# 	msg['Subject'] = "Saarang Freshie orientation | QRCode"
		
	# 	image = MIMEImage(image_read)
	# 	msg.attach(image)
	# 	html =  """\
	# 				<html>
	# 				  <head></head>
	# 				  <body>
	# 				    <p>Greetings from Saarang 2018 team.<br>
	# 				     Entry into CLT will be done through a unique QR code verification.Please find attached your unique QR code.<br>
	# 				     <br>You can either show this on your smartphone or bring a printout of it you don't have a smartphone.<br><br>
	# 				     <b>Also we'll be giving free Saarang t-shirt and keychain on entry upon showing the qr code through a lucky draw.</b>
	# 				     <br>Venue: CLT | Time: 8pm<br><br>
	# 				     For queries contact Uday:+91 80561 97476

	# 				    </p>
	# 				  </body>
	# 				</html>
	# 			"""
	# 	part1 = MIMEText(html, 'html')
	# 	msg.attach(part1)
	# 	s.sendmail(msg['From'], msg['To'], msg.as_string())
	# 	# dictionary = dict()
	# 	# dictionary['to'] = 'rohit@saarang.org'
	# 	# dictionary['from'] = 'webadmin@saarang.org'
	# 	# dictionary['sub'] = 'Saarang Events QRCode'
	# 	# dictionary['html'] = "<h1>This is your QRCode for Saarang events. Please bring this to the event</h1><img src='data:image/jpeg;base64," + str(encodedimage) + "'>"
	# 	# r=requests.post("https://notify.saarang.org/v1/send/email",json = json.loads(json.dumps(dictionary)))
	# 	# print(r.status_code)
	# 	# print(r.json())
	
	# print("\n\n\nBoys over\n\n\n")
	
	# for filename in os.listdir('QRCodes/Girls/'):
	# 	msg = MIMEMultipart()
	# 	msg['From'] =  'do-not-reply@saarang.org'

	# 	body, ext = os.path.splitext(filename)
	# 	emailId=str.lower(str(body))+"@smail.iitm.ac.in"
	# 	imageopen = open(os.path.join('QRCodes/Girls/',filename), 'rb')
	# 	image_read = imageopen.read()
	# 	# encodedimage = encode(image_read)
	# 	# print(str(encodedimage))		
	# 	msg['To'] = emailId
	# 	msg['Subject'] = "Saarang Freshie orientation | QRCode"
	# 	print(msg['To'])
	# 	image = MIMEImage(image_read)
	# 	msg.attach(image)
	# 	html =  """\
	# 				<html>
	# 				  <head></head>
	# 				  <body>
	# 				    <p>Greetings from Saarang 2018 team.<br>
	# 				     Entry into CLT will be done through a unique QR code verification.Please find attached your unique QR code.<br>
	# 				     <br>You can either show this on your smartphone or bring a printout of it you don't have a smartphone.<br><br>
	# 				     <b>Also we'll be giving free Saarang t-shirt and keychain on entry upon showing the qr code through a lucky draw.</b>
	# 				     <br>Venue: CLT | Time: 8pm<br><br>
	# 				     For queries contact Uday:+91 80561 97476

	# 				    </p>
	# 				  </body>
	# 				</html>
	# 			"""
	# 	part1 = MIMEText(html, 'html')
	# 	msg.attach(part1)
	# 	s.sendmail(msg['From'], msg['To'], msg.as_string())
	
	# edm_night_file.close()
	# rock_show_file.close()
	# popular_night_file.close()
	# choreo_night_file.close()
	s.quit()

				

